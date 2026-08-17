import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from sqlalchemy.orm import Session
from backend import models

LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend", "src", "assets", "logo.png")

def generate_title_pdf(title: models.Title, questions_list: list, db: Session) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1e3a8a'), # Indigo/Navy
        alignment=1 # Centered
    )
    
    subtitle_style = ParagraphStyle(
        name='SubtitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#475569'), # Slate
        alignment=1
    )

    body_bold = ParagraphStyle(
        name='BodyBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
        textColor=colors.HexColor('#1e293b')
    )

    table_header_style = ParagraphStyle(
        name='TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
        textColor=colors.white
    )

    table_body_style = ParagraphStyle(
        name='TableBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#334155')
    )

    # Header with SJC Logo
    if os.path.exists(LOGO_PATH):
        try:
            # Scale logo to fit nicely in report
            story.append(Image(LOGO_PATH, width=50, height=85))
            story.append(Spacer(1, 10))
        except Exception:
            pass
            
    story.append(Paragraph("SJC IQAC-IMS (Institutional Monitoring System)", title_style))
    story.append(Spacer(1, 5))
    story.append(Paragraph(f"Title Checklist Report: {title.name}", subtitle_style))
    if title.description:
        story.append(Spacer(1, 5))
        story.append(Paragraph(f"Description: {title.description}", subtitle_style))
    
    story.append(Spacer(1, 15))
    
    # Calc progress
    total = len(questions_list)
    completed = sum(1 for q in questions_list if q.status and q.status.is_checked)
    pct = (completed / total * 100) if total > 0 else 0
    
    progress_text = f"Compliance: {completed}/{total} Questions Checked ({pct:.1f}%) | Exported on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    story.append(Paragraph(progress_text, body_bold))
    story.append(Spacer(1, 10))
    
    # Checklist Table
    data = [[
        Paragraph("S.No", table_header_style),
        Paragraph("Question Description", table_header_style),
        Paragraph("Status", table_header_style),
        Paragraph("Ticked Date & Time", table_header_style),
        Paragraph("Updated By", table_header_style)
    ]]
    
    for idx, q in enumerate(questions_list, 1):
        status_text = "✓ Checked" if q.status and q.status.is_checked else "❌ Pending"
        ticked_time_str = "-"
        updated_by_str = "-"
        
        if q.status and q.status.is_checked:
            if q.status.ticked_at:
                ticked_time_str = q.status.ticked_at.strftime('%Y-%m-%d %I:%M %p')
                if q.status.is_manual_time:
                    ticked_time_str += " (Manual)"
            if q.status.user:
                updated_by_str = q.status.user.username

        data.append([
            Paragraph(str(idx), table_body_style),
            Paragraph(q.text, table_body_style),
            Paragraph(status_text, table_body_style),
            Paragraph(ticked_time_str, table_body_style),
            Paragraph(updated_by_str, table_body_style)
        ])
    
    # Table styling
    col_widths = [30, 230, 75, 120, 75]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f46e5')), # Indigo-600
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
    ])
    
    # Alternating row colors
    for i in range(1, len(data)):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8fafc'))
            
    t.setStyle(t_style)
    story.append(t)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_title_excel(title: models.Title, questions_list: list, db: Session) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist Report"
    
    # Styling variables
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="1E3A8A")
    meta_font = Font(name="Arial", size=10, italic=True, color="475569")
    regular_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border_side = Side(border_style="thin", color="CBD5E1")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Title Blocks
    ws.merge_cells("A1:E1")
    ws["A1"] = "SJC IQAC-IMS (Institutional Monitoring System)"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Title Checklist Report: {title.name}"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center_align
    
    # Progress Meta
    total = len(questions_list)
    completed = sum(1 for q in questions_list if q.status and q.status.is_checked)
    pct = (completed / total * 100) if total > 0 else 0
    
    ws.merge_cells("A3:E3")
    ws["A3"] = f"Compliance Status: {completed}/{total} Questions Checked ({pct:.1f}%) | Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = meta_font
    ws["A3"].alignment = center_align
    
    ws.append([]) # Row 4 empty
    
    # Headers
    headers = ["S.No", "Question Description", "Status", "Ticked Date & Time", "Updated By"]
    ws.append(headers)
    
    # Style header row (Row 5)
    for col_idx in range(1, 6):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Append questions
    for idx, q in enumerate(questions_list, 1):
        status_text = "Checked" if q.status and q.status.is_checked else "Pending"
        ticked_time_str = "-"
        updated_by_str = "-"
        
        if q.status and q.status.is_checked:
            if q.status.ticked_at:
                ticked_time_str = q.status.ticked_at.strftime('%Y-%m-%d %I:%M %p')
                if q.status.is_manual_time:
                    ticked_time_str += " (Manual)"
            if q.status.user:
                updated_by_str = q.status.user.username
                
        row_data = [idx, q.text, status_text, ticked_time_str, updated_by_str]
        ws.append(row_data)
        
        # Style current row
        curr_row = ws.max_row
        for col_idx in range(1, 6):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
                
            # Zebra striping
            if curr_row % 2 == 0:
                cell.fill = alt_row_fill
                
    # Set column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 18
    
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[5].height = 25
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_mails_pdf(mails_list: list) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1e3a8a'),
        alignment=1
    )
    
    subtitle_style = ParagraphStyle(
        name='SubtitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#475569'),
        alignment=1
    )

    body_bold = ParagraphStyle(
        name='BodyBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
        textColor=colors.HexColor('#1e293b')
    )

    table_header_style = ParagraphStyle(
        name='TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )

    table_body_style = ParagraphStyle(
        name='TableBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#334155')
    )

    if os.path.exists(LOGO_PATH):
        try:
            story.append(Image(LOGO_PATH, width=50, height=85))
            story.append(Spacer(1, 10))
        except Exception:
            pass

    story.append(Paragraph("SJC IQAC-IMS (Institutional Monitoring System)", title_style))
    story.append(Spacer(1, 5))
    story.append(Paragraph("Mail Tracking Registry Report", subtitle_style))
    story.append(Spacer(1, 15))
    
    # Progress/Stats
    total = len(mails_list)
    answered = sum(1 for m in mails_list if m.is_answered)
    pending = total - answered
    ans_rate = (answered / total * 100) if total > 0 else 0
    
    meta_text = f"Total Queries: {total} | Answered: {answered} ({ans_rate:.1f}%) | Pending: {pending} | Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    story.append(Paragraph(meta_text, body_bold))
    story.append(Spacer(1, 10))
    
    # Table headers
    data = [[
        Paragraph("S.No", table_header_style),
        Paragraph("Mail Subject / Query", table_header_style),
        Paragraph("Sender Staff", table_header_style),
        Paragraph("Sent Date & Time", table_header_style),
        Paragraph("Status", table_header_style),
        Paragraph("Answered Date & Time", table_header_style)
    ]]
    
    for idx, m in enumerate(mails_list, 1):
        status_text = "✓ Answered" if m.is_answered else "⏳ Waiting"
        sent_str = m.sent_at.strftime('%Y-%m-%d %I:%M %p')
        if m.is_manual_sent_time:
            sent_str += " (M)"
            
        ans_str = "-"
        if m.is_answered and m.answered_at:
            ans_str = m.answered_at.strftime('%Y-%m-%d %I:%M %p')
            if m.is_manual_answered_time:
                ans_str += " (M)"

        data.append([
            Paragraph(str(idx), table_body_style),
            Paragraph(m.subject, table_body_style),
            Paragraph(m.sender_staff, table_body_style),
            Paragraph(sent_str, table_body_style),
            Paragraph(status_text, table_body_style),
            Paragraph(ans_str, table_body_style)
        ])
        
    col_widths = [30, 160, 90, 105, 65, 110]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d9488')), # Teal-600
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
    ])
    
    for i in range(1, len(data)):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8fafc'))
            
    t.setStyle(t_style)
    story.append(t)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_mails_excel(mails_list: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mail Tracking Report"
    
    teal_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="0D9488")
    meta_font = Font(name="Arial", size=10, italic=True, color="475569")
    regular_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border_side = Side(border_style="thin", color="CBD5E1")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.merge_cells("A1:F1")
    ws["A1"] = "SJC IQAC-IMS (Institutional Monitoring System)"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    
    ws.merge_cells("A2:F2")
    ws["A2"] = "Mail Tracking Registry Report"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center_align
    
    total = len(mails_list)
    answered = sum(1 for m in mails_list if m.is_answered)
    pending = total - answered
    ans_rate = (answered / total * 100) if total > 0 else 0
    
    ws.merge_cells("A3:F3")
    ws["A3"] = f"Total Mails: {total} | Answered: {answered} ({ans_rate:.1f}%) | Pending: {pending} | Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = meta_font
    ws["A3"].alignment = center_align
    
    ws.append([]) # Empty row 4
    
    headers = ["S.No", "Mail Subject / Query", "Sender Staff", "Sent Date & Time", "Status", "Answered Date & Time"]
    ws.append(headers)
    
    # Format Headers (Row 5)
    for col_idx in range(1, 7):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = teal_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    for idx, m in enumerate(mails_list, 1):
        status_text = "Answered" if m.is_answered else "Waiting"
        sent_str = m.sent_at.strftime('%Y-%m-%d %I:%M %p')
        if m.is_manual_sent_time:
            sent_str += " (Manual)"
            
        ans_str = "-"
        if m.is_answered and m.answered_at:
            ans_str = m.answered_at.strftime('%Y-%m-%d %I:%M %p')
            if m.is_manual_answered_time:
                ans_str += " (Manual)"
                
        row_data = [idx, m.subject, m.sender_staff, sent_str, status_text, ans_str]
        ws.append(row_data)
        
        curr_row = ws.max_row
        for col_idx in range(1, 7):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [2, 3]:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
                
            if curr_row % 2 == 0:
                cell.fill = alt_row_fill
                
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 22
    
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[5].height = 25
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
